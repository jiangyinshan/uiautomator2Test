from openpyxl import *
import logging

from AutoInput.manage_case_road import manage_case_road


class excel():
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, column, cellvalue):
        try:
            self.ws.cell(row=row, column=column).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=column).value = "保存失败"
            self.wb.save(self.file)

    def assertPass(self, row, column):
        try:
            self.ws.cell(row=row, column=column).value = "通过"
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=column).value = "保存失败"
            self.wb.save(self.file)

    def assertFail(self, row, column):
        try:
            self.ws.cell(row=row, column=column).value = "未通过"
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=column).value = "保存失败"
            self.wb.save(self.file)

    def saveResult(self, coordinate, result):
        try:
            self.ws.cell(row=coordinate[0], column=coordinate[1]).value = result
            self.wb.save(self.file)
        except:
            self.ws.cell(row=coordinate[0], column=coordinate[1]).value = "保存失败"
            self.wb.save(self.file)

    def saveData(self, file):
        self.wb.save(file)

    def equalAssert(self, finalContent, expectContent, conclusion):
        if finalContent.lower().strip() == expectContent.lower().strip():
            self.assertPass(conclusion[0], conclusion[1])
        else:
            self.assertFail(conclusion[0], conclusion[1])

    def containAssert(self, finalContent, expectContent, conclusion):
        if expectContent.lower().strip() in finalContent.lower().strip():
            self.assertPass(conclusion[0], conclusion[1])
        else:
            self.assertFail(conclusion[0], conclusion[1])
